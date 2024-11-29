import joblib
import numpy as np
import pandas as pd
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows
import fasttext
from scipy import spatial
import os
import pickle

from sklearn.neighbors import KNeighborsRegressor

from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from llama_index.core import GPTVectorStoreIndex, download_loader, StorageContext, load_index_from_storage

# Установка переменной окружения для API ключа
os.environ['OPENAI_API_KEY'] = YOUR_KEY

# Авторизация Google Docs
def authorize_gdocs():
    google_oauth2_scopes = [
        "https://www.googleapis.com/auth/documents.readonly"
    ]
    cred = None
    if os.path.exists("token.pickle"):
        with open("token.pickle", 'rb') as token:
            cred = pickle.load(token)
    if not cred or not cred.valid:
        if cred and cred.expired and cred.refresh_token:
            cred.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", google_oauth2_scopes)
            cred = flow.run_local_server(port=0)
        with open("token.pickle", 'wb') as token:
            pickle.dump(cred, token)

def ASK_GPT(que):
    prompt = "Напиши на русском языке в нескольких предложениях ответ на вопрос: " + que
    response = query_engine.query(prompt)
    return response

def read_excel_to_df(filexlsx):
    database = pd.read_excel(filexlsx)
    return database

def saving_to_xlsx(database, filexlsx):
    wb = op.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(database, index=True, header=True):
        ws.append(r)
    ws.delete_rows(2)
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    wb.save(filexlsx)

Path = YOUR_PATH

def load_knn_model(model_path=Path + 'knn_model_euc.pkl'):
    knn = joblib.load(model_path)
    print(f"Model knn loaded from {model_path}")
    return knn

def predict_knn(distance):
    X = np.array(distance).reshape(-1, 1)
    predicted_scores = knn.predict(X)
    rounded_scores = np.round(predicted_scores).astype(int)
    return rounded_scores


def spatialDistance(vector1, vector2):
    return spatial.distance.euclidean(vector1, vector2)


ft = fasttext.load_model(Path + 'cc.ru.300.bin')
print('Модель загрузилась')

def no_n(text):
    text = str(text)
    while '\n' in text:
        text = text.replace('\n', ' ')
    return text

def generateVector(sentence):
    return ft.get_sentence_vector(no_n(sentence))

def predictions(filexlsx_path, knnmodel_path, index_path):
    # Шаг 1: Загрузка данных
    df = read_excel_to_df(filexlsx_path)
    # Создать ГПТшные ответы:
    storage_context = StorageContext.from_defaults(persist_dir=index_path)
    index = load_index_from_storage(storage_context)
    query_engine = index.as_query_engine()
    df['answ2'] = df['que'].apply(ASK_GPT)
    # Векторизовать ответы:
    df['vector'] = df['answ'].apply(generateVector)
    df['vector2'] = df['answ2'].apply(generateVector)
    print('Первичные векторы сгенерированы')

    #Сравнить один текст с другим:
    df['dist'] = df.apply(lambda x: spatialDistance(x['vector'], x['vector2']), axis=1)
    print(df)
    print('Сравнили векторы')

    knn = load_knn_model(knnmodel_path)
    # Шаг 6: Загрузка модели из файла pkl
    loaded_knn = load_knn_model
    s = df['dist'].to_numpy()

    df['predicted'] = predict_knn(s)
    print('kNN predicted')

    df.drop(columns=['vector', 'vector2'], inplace=True)
    saving_to_xlsx(df, filexlsx_path)
    print('Файл сохранён')

predictions(modelPath+'20.07_add_base.xlsx', modelPath + 'knn_model_euc.pkl', modelPath + "index.json")