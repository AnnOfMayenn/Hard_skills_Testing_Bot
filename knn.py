import os
import pickle
import random
import telebot
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows
import fasttext
import pandas as pd

import joblib
import numpy as np
from sklearn.neighbors import KNeighborsRegressor


from scipy.spatial import distance
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from llama_index.core import GPTVectorStoreIndex, download_loader, StorageContext, load_index_from_storage


Path = YOUR_PATH

# Чтение файла Excel в DataFrame
def saving_to_xlsx(database, filexlsx):
    wb = op.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(database, index=True, header=True):
        ws.append(r)
    ws.delete_rows(2)
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    wb.save(filexlsx)

def read_excel_to_df(filexlsx):
    database = pd.read_excel(filexlsx, index_col=0)
    return database

def save_ques(df2, path=Path + 'que_base.xlsx'):
    df = read_excel_to_df(path)
    print(df)
    print(df2)
    newdf = pd.concat([df, df2], ignore_index=True)
    saving_to_xlsx(newdf, path)

def read_users():
    df = read_excel_to_df(Path + 'user_base.xlsx')
    return df

def save_users():
    saving_to_xlsx(users, Path + 'user_base.xlsx')

users = read_users()
que = pd.DataFrame(columns=['user_id', 'que', 'answ_user', 'answ_gpt', 'score'])

Path = YOUR_PATH
ft = fasttext.load_model(Path + 'cc.ru.300.bin')
print('Модель фасттекста загрузилась')

def load_knn_model(model_path=Path + '/knn_model.pkl'):
    knn = joblib.load(model_path)
    print(f"Model knn loaded from {model_path}")
    return knn

def predict_knn(distance, knn):
    X = np.array(distance).reshape(-1, 1)
    predicted_scores = knn.predict(X)
    rounded_scores = np.round(predicted_scores).astype(int)
    return rounded_scores

# Загрузка модели
knn_model = load_knn_model()

# Установка переменной окружения для API ключа
os.environ['OPENAI_API_KEY'] = YOUR_KEY

# Авторизация Google Docs
def authorize_gdocs():
    google_oauth2_scopes = [
        "https://www.googleapis.com/auth/documents.readonly"
    ]
    cred = None
    if os.path.exists("token.pickle"):
        with open("token.pickle", 'rb') as token:
            cred = pickle.load(token)
    if not cred or not cred.valid:
        if cred and cred.expired and cred.refresh_token:
            cred.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", google_oauth2_scopes)
            cred = flow.run_local_server(port=0)
        with open("token.pickle", 'wb') as token:
            pickle.dump(cred, token)

# Функция для получения вектора текста
def get_text_vector(text):
    return ft.get_sentence_vector(no_n(text))

def no_n(text):
    text = str(text)
    while '\n' in text:
        text = text.replace('\n', ' ')
    print('Текст очищен от \\n')
    return text

# Функция для сравнения двух векторов по расстоянию чебышёва
def spatialDistance(vector1, vector2):
    return distance.euclidean(vector1, vector2)

# Директория для хранения индекса
dir = Path + "index.json"
# Функция для создания новых индексов
def NEW_INDEXES():
    # Инициализация LlamaIndex Google Docs reader
    GoogleDocsReader = download_loader('GoogleDocsReader')
    gdoc_ids = [YOUR_GOOGLE_DOC_LINK]
    loader = GoogleDocsReader()
    authorize_gdocs()
    # Загрузка и индексирование документов
    documents = loader.load_data(document_ids=gdoc_ids)
    index = GPTVectorStoreIndex(documents)

    # Сохранение индекса
    index.storage_context.persist(persist_dir=dir)
    print('Индексы были обновлены')

# Загрузка индекса из сохраненного файла
storage_context = StorageContext.from_defaults(persist_dir=dir)
index = load_index_from_storage(storage_context)
query_engine = index.as_query_engine()

user_direction = {}

# Бот
TOKEN = YOUR_TOKEN
bot = telebot.TeleBot(TOKEN)
print('Запускаем бот')

# Список вопросов и ответов
STACK = {"Frontend": ['Что такое React JS?', 'Какие сильные стороны React JS?', 'Что такое JSX?', 'Что является компонентом в React?', 'За счет чего в React JS обеспечивается высокая производительность?', 'Что такое virtual dom и зачем он нужен?', 'Как работает  Virtual DOM?', 'Как React JS реализован алгоритм согласования Virtual Dom c реальным DOM (Reconcilation)?', 'Для чего нужен атрибут key в компоненте и когда его имеет смысл использовать?', 'Какие действия вызывают перерендер React компонента?', 'Как можно избежать лишних перерендеров React компонента?'],
    "Backend": ['Различие между array и list в python, какие плюсы одного и другого?', 'Как работает алгоритм подсчета ссылок в Garbage Collector?',
                'Различие classmethod и staticmethod в python? Когда стоит применять один а когда другой?', 'В чем суть CAP теоремы?', 'Основное преимущество JWT над сессионой авторизацией?',
                'Как происходит и для чего нужна процедура VACUUM в PostgreSQL?', 'Для чего нужны atomic операции в Django?', 'В чем различия code-first и schema first подходов при описании GraphQL схемы? В каких случаях может быть полезен schema first подход?', 'Что такое контейнеризация, какой механизм изоляции в ней используется?',
                'Отличие протоколов RPC и REST?'],
    "C++": ["Чем полезны умные указатели при работе с памятью?",
            "Что такое RAII (Resource Acquisition Is Initialization) и зачем он нужен?", "Зачем нужен RTTI (Runtime type identification)?", 'Что такое move семантика, и где она применяется?',
            'Как работает множественное наследование в C++, и какие проблемы могут возникнуть при его использовании?',
            'Чем отличается std::async от std::thread в STL?', 'Что такое CRTP (Curiously Recurring Template Pattern) в C++?',
            'Что может быть не так с порядком инициализации/деинициализации статических объектов в С++?', 'Что такое фрагментация памяти, и почему важно помнить о ней?',
            'Чем примечательны COW (Copy-On-Write) и SSO (Small String Optimization) строки в С++, в каких случаях преимущество имеют одни а в каких другие?']}

# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row('Начать')
    bot.send_message(message.chat.id, "Добро пожаловать в 3DiVi hard-skills testing Bot", reply_markup=markup)

# Обработчик кнопки "Сменить направление"
@bot.message_handler(func=lambda message: message.text in ['Сменить направление', 'Начать'])
def choose_direction(message):
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row('Frontend', 'Backend', 'C++')
    bot.send_message(message.chat.id, "Выберите направление", reply_markup=markup)

# Обработчик кнопок "Frontend", "Backend", "C++"
@bot.message_handler(func=lambda message: message.text in ['Frontend', 'Backend', 'C++'])
def change_direction(message):
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row('Начать тестирование', 'Сменить направление', 'Обновить информацию')
    user_id = message.from_user.id
    user_name = message.from_user.first_name
    if user_id in users['user_id'].values:
        users.loc[users['user_id'] == user_id, 'stack'] = message.text
    else:
        users.loc[len(users.index)] = [user_id, user_name, message.text]
    print(users)
    bot.send_message(message.chat.id, f"Выбрано направление: {message.text}", reply_markup=markup)

# Обработчик кнопки "Обновить информацию"
@bot.message_handler(func=lambda message: message.text == 'Обновить информацию')
def update_info(message):
    bot.send_message(message.chat.id, "Введите код:")
    bot.register_next_step_handler(message, process_code)

def process_code(message):
    if message.text == "12345":
        bot.send_message(message.chat.id, YOUR_GOOGLE_DOC)
        markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add('Файл отредактирован')
        bot.send_message(message.chat.id, "Пожалуйста, нажмите кнопку, когда вы закончите редактировать файл", reply_markup=markup)
    else:
        markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row('Начать тестирование', 'Сменить направление', 'Обновить информацию')
        bot.send_message(message.chat.id, "Неверный код. В доступе отказано", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == 'Файл отредактирован')
def renew_indexes(message):
    NEW_INDEXES()
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row('Начать тестирование', 'Сменить направление', 'Обновить информацию')
    bot.send_message(message.chat.id, "Успешно отредактирован файл!", reply_markup=markup)

# Обработчик кнопки "Начать тестирование"
@bot.message_handler(func=lambda message: message.text in ['Начать тестирование', 'Продолжить'])
def start_testing(message):
    user_id = message.from_user.id
    question = random.choice(STACK[users.loc[users['user_id'] == user_id, 'stack'].values[0]])
    user_direction[user_id] = question
    bot.send_message(message.chat.id, f"Вопрос: {question}")
    bot.register_next_step_handler(message, handle_answer)

def ASK_GPT(que):
    prompt = "Напиши на русском языке в нескольких предложениях ответ на вопрос: " + que
    response = query_engine.query(prompt)
    return response

# Обработчик ответов на вопросы
def handle_answer(message):
    user_id = message.from_user.id
    answ_user = message.text
    if 'DFsave' in answ_user:
        markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row('Продолжить', 'Закончить')
        save_users()
        save_ques(que)
        bot.send_message(message.chat.id,
                         "Локальные базы данных с вопросами и с юзерами сохранены",
                         reply_markup=markup)

    else:
        answ_gpt = ASK_GPT(user_direction[user_id])
        vector_user = get_text_vector(answ_user)
        vector_gpt = get_text_vector(answ_gpt)
        similarity_score = spatialDistance(vector_user, vector_gpt)
        predicted_scores = predict_knn(similarity_score, knn_model)[0]
        que.loc[len(que.index)] = [user_id, user_direction[user_id], answ_user, answ_gpt, predicted_scores]
        user_direction.pop(user_id, None)
        markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row('Продолжить', 'Закончить')
        bot.send_message(message.chat.id,
                             f"Ваш ответ: {answ_user}. \n\nЭкспертный ответ: {answ_gpt} \n\nОценка: {predicted_scores} \n\n Чтобы завершить тест, нажмите на кнопку \"Закончить\". \n Чтобы продолжить тестирование, нажмите на кнопку \"Продолжить\"",
                             reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == 'Закончить')
def main(message):
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    # Запуск бота
    markup.row('Начать тестирование', 'Сменить направление', 'Обновить информацию')
    bot.send_message(message.chat.id, "Спасибо за прохождение теста!", reply_markup=markup)

bot.infinity_polling(timeout=10, long_polling_timeout = 5)
